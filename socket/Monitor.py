import openpyxl
import os
import ServerOperation
from ServerOperation import ServerData
import threading
from threading import Thread

AuthorityInfoPath = "Info\\Authority.xlsx" # 存放管理员和经理密码的表格

class Manager():

    def __init__(self):
        run = ServerData()
        MonitorThread = Thread(target=self.FirstUIDeal)
        MonitorThread.setDaemon(True)
        MonitorThread.start()
        run.SocketStart()


    def LogNameCheck(self, id, password):  # 检查用户登录信息是否正确
        isExist = os.path.exists(AuthorityInfoPath)  # 检查登记表是否存在
        if isExist != 1:  # 还没有在主机端创建管理员信息表，就直接创建一个新的表
            os.makedirs("Info")
            table = openpyxl.Workbook()
            table.save(AuthorityInfoPath)
        excel = openpyxl.load_workbook(AuthorityInfoPath)
        data = excel.worksheets[0]
        #  passwordForCheck = data.cell(id - 1, 1).value  # 获取对应身份的密码
        passwordForCheck = "12345"  # 这里为了测试方便就直接用固定值
        print("###########",password)
        if id == 1 and password == passwordForCheck:
            print("Correct...")
            return True
        else:
            print("The password is WRONG...")
            return False

    def PasswordChange(self, id, password):  # 修改管理员账户密码
        answer = self.LogNameCheck(id, password)
        if answer == False:
            return -1

        excel = openpyxl.load_workbook(AuthorityInfoPath)
        data = excel.worksheets[0]
        FirstPassword = input("Please input your new password :")
        SecondPassword = input("Please input your new password for second time:")
        if FirstPassword == SecondPassword:
            print("Change successful!...")
            # 修改管理员账户的密码

            data.cell(id - 1, 1, FirstPassword)
            excel.save(AuthorityInfoPath)
            return 0
        else:
            print("The new password is not same,Please Try again...")
            return -1

    def FirstUIDeal(self):  # 登录页，返回Flag表明输入的情况，1为成功，0为退出
        while True:  # 只要没有选择退出或者登陆成功，就一直在登录页面循环
            symbol = input("Choose your Identification: 1 for authority, 2 for header\n")
            #  print("symbol:", symbol)
            if symbol == "1":  # 管理员登录
                password = input("Please input the Password:")
                if self.LogNameCheck(1, password) == True:
                    print("登录成功，开始跳转...")
                    Flag = 1
                    break
            elif symbol == "2":  # 经理登录
                password = input("Please input the password:")
                if self.LogNameCheck(2, password) == True:
                    print("登录成功，开始跳转...")
                    Flag = 1
                    break
            else:  # 选择退出
                Flag = 0
                break
        return Flag

test = Manager()